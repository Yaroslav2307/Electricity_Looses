# Кусок кода по вколючению записи в формулу OPen Document
# import odf
# from odf.opendocument import OpenDocumentText
# from odf.element import Element
# from odf.text import P
# from odf.math import Math
# from namespaces import MATHNS
#
#
# def main():
#     doc = OpenDocumentText()
#     p = P(text=u'text')
#     df = odf.draw.Frame( zindex=0, anchortype='as-char')
#     p.addElement(df)
#     doc.text.addElement(p)
#
#     formula =u'c=sqrt(a^2+b^2)'
#     math = Math()
#     annot = Element(qname = (MATHNS,u'annotation'))
#     annot.addText(formula, check_grammar=False)
#     annot.setAttribute((MATHNS,'encoding'), 'StarMath 5.0', check_grammar=False)
#     math.addElement(annot)
#     do = odf.draw.Object()
#     do.addElement(math)
#     df.addElement(do)
#
#     outputfile = u'result'
#     doc.save(outputfile, True)
#
# if __name__ == '__main__':
#     main()
# Проработка функций по извлечению данных из Excel
from openpyxl import load_workbook  # Подключение функции загрузки рабочей книги
wb = load_workbook('Origin.xlsx')   # Загрузка таблицы из директории файла
print(wb.get_sheet_names())         # Выведение наименований листов
sheet = wb.get_sheet_by_name('Points of measurement')   # Выбор активного листа
print(sheet.title)                                      # Печать наименования активного листа
print(sheet['A1'].value)                                # вывод значения ячейки
c = sheet['B2']
print(c)                                                # вывод адреса значения ячейки - НЕ ЗНАЧЕНИЯ
print(sheet.cell(row=2, column=4).value)                # другой формат вывода значения ячейки
for i in range(1, 4):
     print(i, sheet.cell(row=i, column=2).value)        # печать второго столбца таблицы